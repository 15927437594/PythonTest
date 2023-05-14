import os

from openpyxl import Workbook

path = os.path.join(os.getcwd(), 'log_LiX01_TPCM')

sn_list = []
for root, _, files in os.walk(path):
    for file_name in files:
        print(root, file_name)
        file_path = os.path.join(root, file_name)
        date = file_name.split('.')[0]
        sheet_name = root.split('\\')[1]
        print(file_path, sheet_name)
        with open(file_path, encoding='utf8') as file:
            workbook = Workbook()
            excel_name = 'LiX01_TPCM-%s.xlsx' % date
            tpcm_excel = os.path.join(os.getcwd(), 'export_LiX01_TPCM', excel_name)
            sheet_count = 0
            lines = file.readlines()
            for index in range(len(lines)):
                if lines[index].__contains__("handleTpCmValueDetect cmValue="):
                    sheet_count += 1
                    workbook.create_sheet("Sheet%s" % sheet_count)
                    sheet_item = "Sheet%s" % sheet_count
                    print(sheet_item)
                    sheet = workbook[sheet_item]

                    cm_data = lines[index].split("handleTpCmValueDetect cmValue=")[1]
                    print(cm_data)
                    cm_data = cm_data.replace("\n", "")
                    cm_list = cm_data.split(" ")
                    print(cm_list)

                    # filename = os.path.join(path, 'TPCM.txt')
                    # with open(filename, 'a') as file_object:
                    #
                    #     file_object.write(cm_data)
                    #     file_object.write("\n")

                    for i in range(35):
                        for j in range(62):
                            data = cm_list[62 * i + j]
                            # print(data)
                            sheet.cell(i + 1, j + 1, data)

            workbook.save(tpcm_excel)
